"""
This module provides a toolbox of independent, stateless functions for processing Excel files.
Each function is designed to be a single, robust, and testable unit of work.
"""

import traceback
import os
import time
from typing import Dict, Any, List
from openpyxl import load_workbook
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np

# --- Helper Functions ---

def _get_dataframe(file_path: str, sheet_name: str = None) -> pd.DataFrame:
    """Helper to read an Excel file and return a single DataFrame."""
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    if isinstance(df, dict):
        return list(df.values())[0]
    return df

def _get_safe_path(output_dir: str, filename: str) -> str:
    """Joins an output directory and filename, with basic security checks."""
    if '..' in filename or os.path.isabs(filename):
        raise ValueError(f"Invalid filename '{filename}'. Must be a relative path with no directory traversal.")
    return os.path.join(output_dir, filename)

# --- Read-only Tools ---

def get_data_summary(file_path: str, sheet_name: str = None) -> Dict[str, Any]:
    """Reads an Excel file and returns a summary of its contents without loading data."""
    try:
        workbook = load_workbook(filename=file_path, read_only=True)
        if sheet_name and sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        header_columns = [cell.value for cell in sheet[1]]
        summary = {"sheet_name": sheet.title, "total_rows": sheet.max_row, "header_columns": header_columns}
        workbook.close()
        return summary
    except Exception as e:
        raise Exception(f"An unexpected error occurred while getting data summary: {e}")

def read_rows(file_path: str, sheet_name: str = None, offset: int = 1, limit: int = 5) -> List[Dict[str, Any]]:
    """Reads a specific range of rows from an Excel sheet."""
    try:
        workbook = load_workbook(filename=file_path, read_only=True)
        if sheet_name and sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}")
        sheet = workbook[sheet_name] if sheet_name else workbook.active
        header = [cell.value for cell in sheet[1]]
        data_rows = []
        min_row_to_read = offset + 1
        max_row_to_read = min_row_to_read + limit - 1
        for row in sheet.iter_rows(min_row=min_row_to_read, max_row=max_row_to_read, values_only=True):
            data_rows.append(dict(zip(header, row)))
        workbook.close()
        return data_rows
    except Exception as e:
        raise Exception(f"An unexpected error occurred while reading rows: {e}")

def column_aggregate(file_path: str, column_name: str, aggregate_function: str, sheet_name: str = None) -> Dict[str, Any]:
    """Performs a simple aggregation on a single column."""
    try:
        df = _get_dataframe(file_path, sheet_name)
        if column_name not in df.columns: raise ValueError(f"Column '{column_name}' not found.")
        values = pd.to_numeric(df[column_name], errors='coerce').dropna()
        if values.empty: return {"result": None, "processed_rows": 0}
        result = 0
        if aggregate_function == 'sum': result = values.sum()
        elif aggregate_function == 'mean': result = values.mean()
        elif aggregate_function == 'min': result = values.min()
        elif aggregate_function == 'max': result = values.max()
        return {"result": float(result), "processed_rows": len(values)}
    except Exception as e:
        raise Exception(f"An unexpected error occurred during aggregation: {e}")

def get_unique_values(file_path: str, column_name: str, sheet_name: str = None) -> Dict[str, Any]:
    """Gets a list of unique values from a specified column."""
    try:
        df = _get_dataframe(file_path, sheet_name)
        if column_name not in df.columns: raise ValueError(f"Column '{column_name}' not found.")
        unique_values = df[column_name].unique().tolist()
        return {"unique_values": unique_values, "count": len(unique_values)}
    except Exception as e:
        raise Exception(f"An unexpected error occurred while getting unique values: {e}")

# --- Write Tools (accept output_dir) ---

def add_column_from_formula(file_path: str, file_output_dir: str, output_filename: str, new_column_name: str, col1: str, operator: str, col2: str, sheet_name: str = None) -> Dict[str, Any]:
    """Adds a new column based on a formula and saves to a new file in a safe directory."""
    final_output_path = _get_safe_path(file_output_dir, output_filename)
    try:
        df = _get_dataframe(file_path, sheet_name)
        if col1 not in df.columns or col2 not in df.columns: raise ValueError(f"Column not found.")
        val1 = pd.to_numeric(df[col1], errors='coerce')
        val2 = pd.to_numeric(df[col2], errors='coerce')
        if operator == '+': df[new_column_name] = val1 + val2
        elif operator == '-': df[new_column_name] = val1 - val2
        elif operator == '*': df[new_column_name] = val1 * val2
        elif operator == '/': df[new_column_name] = val1 / val2
        df.to_excel(final_output_path, index=False)
        return {"success": True, "output_file": final_output_path}
    except Exception as e:
        raise Exception(f"An unexpected error occurred while adding column: {e}")

def create_chart(file_path: str, chart_output_dir: str, chart_type: str, x_column: str, y_column: str, output_filename: str = None, sheet_name: str = None) -> Dict[str, Any]:
    """Creates a chart and saves it as an image file in a safe directory."""
    if output_filename is None:
        timestamp = int(time.time())
        output_filename = f"chart_{timestamp}.png"
    final_output_path = _get_safe_path(chart_output_dir, output_filename)
    try:
        df = _get_dataframe(file_path, sheet_name)
        if x_column not in df.columns or y_column not in df.columns: raise ValueError(f"Column not found.")
        plt.figure(figsize=(10, 6))
        if chart_type == 'bar': plt.bar(df[x_column], df[y_column])
        elif chart_type == 'line': plt.plot(df[x_column], df[y_column])
        elif chart_type == 'scatter': plt.scatter(df[x_column], df[y_column])
        plt.xlabel(x_column); plt.ylabel(y_column); plt.title(f'{y_column} vs. {x_column}'); plt.grid(True)
        plt.savefig(final_output_path)
        plt.close()
        return {"success": True, "chart_path": final_output_path}
    except Exception as e:
        raise Exception(f"An unexpected error occurred while creating chart: {e}")

def filter_rows(file_path: str, file_output_dir: str, output_filename: str, column_name: str, operator: str, value: Any, sheet_name: str = None) -> Dict[str, Any]:
    """Filters rows based on a condition and saves the result to a new file in a safe directory."""
    final_output_path = _get_safe_path(file_output_dir, output_filename)
    try:
        df = _get_dataframe(file_path, sheet_name)
        if column_name not in df.columns: raise ValueError(f"Column not found.")
        column_type = df[column_name].dtype
        try:
            if pd.api.types.is_numeric_dtype(column_type): value = float(value)
        except (ValueError, TypeError): pass
        if operator == '==': filtered_df = df[df[column_name] == value]
        elif operator == '!=': filtered_df = df[df[column_name] != value]
        elif operator == '>': filtered_df = df[df[column_name] > value]
        elif operator == '<': filtered_df = df[df[column_name] < value]
        elif operator == '>=': filtered_df = df[df[column_name] >= value]
        elif operator == '<=': filtered_df = df[df[column_name] <= value]
        elif operator == 'contains': filtered_df = df[df[column_name].astype(str).str.contains(str(value))]
        else: raise ValueError(f"Operator {operator} not handled.")
        filtered_df.to_excel(final_output_path, index=False)
        return {"success": True, "output_file": final_output_path, "rows_written": len(filtered_df)}
    except Exception as e:
        raise Exception(f"An unexpected error occurred while filtering rows: {e}")

def sort_data(file_path: str, file_output_dir: str, output_filename: str, sort_by_column: str, ascending: bool = True, sheet_name: str = None) -> Dict[str, Any]:
    """Sorts the data by a specific column and saves to a new file in a safe directory."""
    final_output_path = _get_safe_path(file_output_dir, output_filename)
    try:
        df = _get_dataframe(file_path, sheet_name)
        if sort_by_column not in df.columns: raise ValueError(f"Column not found.")
        sorted_df = df.sort_values(by=sort_by_column, ascending=ascending)
        sorted_df.to_excel(final_output_path, index=False)
        return {"success": True, "output_file": final_output_path}
    except Exception as e:
        raise Exception(f"An unexpected error occurred while sorting data: {e}")

def create_pivot_table(file_path: str, file_output_dir: str, output_filename: str, index_column: str, columns_column: str, values_column: str, agg_func: str = 'sum', sheet_name: str = None) -> Dict[str, Any]:
    """Creates a pivot table and saves it to a new file in a safe directory."""
    final_output_path = _get_safe_path(file_output_dir, output_filename)
    try:
        df = _get_dataframe(file_path, sheet_name)
        pivot_df = df.pivot_table(index=index_column, columns=columns_column, values=values_column, aggfunc=agg_func)
        pivot_df.to_excel(final_output_path)
        return {"success": True, "output_file": final_output_path}
    except Exception as e:
        raise Exception(f"An unexpected error occurred while creating pivot table: {e}")

def delete_columns(file_path: str, file_output_dir: str, output_filename: str, columns_to_delete: List[str], sheet_name: str = None) -> Dict[str, Any]:
    """Deletes one or more columns from a sheet and saves to a new file."""
    final_output_path = _get_safe_path(file_output_dir, output_filename)
    try:
        df = _get_dataframe(file_path, sheet_name)
        original_columns = df.columns.tolist()
        
        # Check for columns that exist in the dataframe
        existing_columns_to_delete = [col for col in columns_to_delete if col in df.columns]
        if not existing_columns_to_delete:
            raise ValueError(f"None of the specified columns {columns_to_delete} exist in the file.")

        df.drop(columns=existing_columns_to_delete, inplace=True)
        
        df.to_excel(final_output_path, index=False)
        
        return {"success": True, "output_file": final_output_path, "deleted_columns": existing_columns_to_delete}
    except Exception as e:
        raise Exception(f"An unexpected error occurred while deleting columns: {e}")

def rename_column(file_path: str, file_output_dir: str, output_filename: str, old_column_name: str, new_column_name: str, sheet_name: str = None) -> Dict[str, Any]:
    """Renames a single column and saves the result to a new file."""
    final_output_path = _get_safe_path(file_output_dir, output_filename)
    try:
        df = _get_dataframe(file_path, sheet_name)
        if old_column_name not in df.columns:
            raise ValueError(f"Column '{old_column_name}' not found in the file.")
        
        df.rename(columns={old_column_name: new_column_name}, inplace=True)
        
        df.to_excel(final_output_path, index=False)
        
        return {"success": True, "output_file": final_output_path, "renamed_from": old_column_name, "renamed_to": new_column_name}
    except Exception as e:
        raise Exception(f"An unexpected error occurred while renaming column: {e}")

def handle_duplicates(file_path: str, file_output_dir: str, output_filename: str, subset_columns: List[str] = None, action: str = 'find', sheet_name: str = None) -> Dict[str, Any]:
    """Finds or removes duplicate rows based on a subset of columns and saves the result."""
    final_output_path = _get_safe_path(file_output_dir, output_filename)
    try:
        df = _get_dataframe(file_path, sheet_name)
        
        # Find duplicates
        duplicates = df[df.duplicated(subset=subset_columns, keep=False)]
        num_duplicates = len(duplicates)

        if action == 'find':
            return {"success": True, "action": "find", "duplicates_found": num_duplicates, "duplicate_rows": duplicates.to_dict('records')}
        
        elif action == 'remove':
            df.drop_duplicates(subset=subset_columns, keep='first', inplace=True)
            df.to_excel(final_output_path, index=False)
            return {"success": True, "action": "remove", "output_file": final_output_path, "duplicates_removed": num_duplicates}
        
        else:
            raise ValueError(f"Invalid action '{action}'. Must be 'find' or 'remove'.")
            
    except Exception as e:
        raise Exception(f"An unexpected error occurred while handling duplicates: {e}")

def fill_missing_values(file_path: str, file_output_dir: str, output_filename: str, column_name: str, fill_value: Any, sheet_name: str = None) -> Dict[str, Any]:
    """Fills missing values (NaN) in a specified column with a given value."""
    final_output_path = _get_safe_path(file_output_dir, output_filename)
    try:
        df = _get_dataframe(file_path, sheet_name)
        if column_name not in df.columns:
            raise ValueError(f"Column '{column_name}' not found in the file.")

        # Attempt to convert fill_value to the column's type if it's numeric
        if pd.api.types.is_numeric_dtype(df[column_name].dtype):
            try:
                fill_value = pd.to_numeric(fill_value)
            except (ValueError, TypeError):
                raise ValueError(f"Cannot convert fill_value '{fill_value}' to numeric type for column '{column_name}'.")

        df[column_name].fillna(fill_value, inplace=True)
        df.to_excel(final_output_path, index=False)

        return {"success": True, "output_file": final_output_path, "filled_column": column_name}
    except Exception as e:
        raise Exception(f"An unexpected error occurred while filling missing values: {e}")

def string_manipulation_in_column(file_path: str, file_output_dir: str, output_filename: str, column_name: str, operation: str, sheet_name: str = None) -> Dict[str, Any]:
    """Performs a string operation (uppercase, lowercase, trim) on a column."""
    final_output_path = _get_safe_path(file_output_dir, output_filename)
    try:
        df = _get_dataframe(file_path, sheet_name)
        if column_name not in df.columns:
            raise ValueError(f"Column '{column_name}' not found in the file.")

        # Ensure the column is of string type for string operations
        df[column_name] = df[column_name].astype(str)

        if operation == 'uppercase':
            df[column_name] = df[column_name].str.upper()
        elif operation == 'lowercase':
            df[column_name] = df[column_name].str.lower()
        elif operation == 'trim':
            df[column_name] = df[column_name].str.strip()
        else:
            raise ValueError(f"Invalid string operation '{operation}'.")

        df.to_excel(final_output_path, index=False)

        return {"success": True, "output_file": final_output_path, "manipulated_column": column_name, "operation": operation}
    except Exception as e:
        raise Exception(f"An unexpected error occurred during string manipulation: {e}")

def list_sheets(file_path: str) -> Dict[str, Any]:
    """Lists all the sheet names in the given Excel workbook."""
    try:
        workbook = load_workbook(filename=file_path, read_only=True)
        sheet_names = workbook.sheetnames
        workbook.close()
        return {"success": True, "sheet_names": sheet_names}
    except Exception as e:
        raise Exception(f"An unexpected error occurred while listing sheets: {e}")

def lookup_and_merge_columns(
    file_path: str,
    file_output_dir: str,
    output_filename: str,
    left_on_column: str,
    right_file_path: str,
    right_on_column: str,
    columns_to_merge: List[str],
    sheet_name: str = None,
    right_sheet_name: str = None
) -> Dict[str, Any]:
    """Performs a VLOOKUP-like operation, merging columns from a second file/sheet based on a key column."""
    final_output_path = _get_safe_path(file_output_dir, output_filename)
    try:
        # Read the left and right dataframes
        left_df = _get_dataframe(file_path, sheet_name)
        right_df = _get_dataframe(right_file_path, right_sheet_name)

        # --- Validation ---
        if left_on_column not in left_df.columns:
            raise ValueError(f"Left key column '{left_on_column}' not found in the primary file.")
        if right_on_column not in right_df.columns:
            raise ValueError(f"Right key column '{right_on_column}' not found in the secondary file.")
        for col in columns_to_merge:
            if col not in right_df.columns:
                raise ValueError(f"Column to merge '{col}' not found in the secondary file.")

        # Perform the merge (equivalent to a left join)
        merged_df = pd.merge(
            left_df,
            right_df[[right_on_column] + columns_to_merge],
            left_on=left_on_column,
            right_on=right_on_column,
            how='left'
        )

        merged_df.to_excel(final_output_path, index=False)

        return {"success": True, "output_file": final_output_path, "rows_written": len(merged_df)}
    except Exception as e:
        raise Exception(f"An unexpected error occurred during the merge operation: {e}")

def group_by_and_aggregate(
    file_path: str,
    file_output_dir: str,
    output_filename: str,
    group_by_column: str,
    agg_column: str,
    agg_functions: List[str],
    sheet_name: str = None
) -> Dict[str, Any]:
    """Groups data by a specified column and performs one or more aggregations on another column."""
    final_output_path = _get_safe_path(file_output_dir, output_filename)
    try:
        df = _get_dataframe(file_path, sheet_name)

        # --- Validation ---
        if group_by_column not in df.columns:
            raise ValueError(f"Group by column '{group_by_column}' not found.")
        if agg_column not in df.columns:
            raise ValueError(f"Aggregation column '{agg_column}' not found.")
        
        # Ensure the aggregation column is numeric
        df[agg_column] = pd.to_numeric(df[agg_column], errors='coerce')

        # Perform the group by and aggregation
        grouped_df = df.groupby(group_by_column)[agg_column].agg(agg_functions).reset_index()

        grouped_df.to_excel(final_output_path, index=False)

        return {"success": True, "output_file": final_output_path, "rows_written": len(grouped_df)}
    except Exception as e:
        raise Exception(f"An unexpected error occurred during the group by operation: {e}")

def conditional_value_column(
    file_path: str,
    file_output_dir: str,
    output_filename: str,
    new_column_name: str,
    source_column: str,
    operator: str,
    value: Any,
    true_value: Any,
    false_value: Any,
    sheet_name: str = None
) -> Dict[str, Any]:
    """Creates a new column with values based on a condition applied to a source column."""
    final_output_path = _get_safe_path(file_output_dir, output_filename)
    try:
        df = _get_dataframe(file_path, sheet_name)
        if source_column not in df.columns:
            raise ValueError(f"Source column '{source_column}' not found.")

        # Get the boolean series from the condition
        column_type = df[source_column].dtype
        try:
            if pd.api.types.is_numeric_dtype(column_type):
                value = float(value)
        except (ValueError, TypeError):
            pass # Keep value as string for comparison
        
        conditions = {
            '==': df[source_column] == value,
            '!=': df[source_column] != value,
            '>': df[source_column] > value,
            '<': df[source_column] < value,
            '>=': df[source_column] >= value,
            '<=': df[source_column] <= value,
            'contains': df[source_column].astype(str).str.contains(str(value))
        }
        
        if operator not in conditions:
            raise ValueError(f"Invalid operator '{operator}'.")

        condition_series = conditions[operator]

        # Create the new column
        df[new_column_name] = np.where(condition_series, true_value, false_value)

        df.to_excel(final_output_path, index=False)

        return {"success": True, "output_file": final_output_path, "rows_written": len(df)}
    except Exception as e:
        raise Exception(f"An unexpected error occurred while creating conditional column: {e}")

def create_sheet(file_path: str, file_output_dir: str, output_filename: str, new_sheet_name: str) -> Dict[str, Any]:
    """Creates a new, empty sheet in a workbook and saves it to a new file."""
    final_output_path = _get_safe_path(file_output_dir, output_filename)
    try:
        workbook = load_workbook(filename=file_path)
        if new_sheet_name in workbook.sheetnames:
            raise ValueError(f"A sheet named '{new_sheet_name}' already exists.")
        
        workbook.create_sheet(title=new_sheet_name)
        workbook.save(final_output_path)
        workbook.close()
        
        return {"success": True, "output_file": final_output_path, "created_sheet_name": new_sheet_name}
    except Exception as e:
        raise Exception(f"An unexpected error occurred while creating the sheet: {e}")

def delete_sheet(file_path: str, file_output_dir: str, output_filename: str, sheet_to_delete: str) -> Dict[str, Any]:
    """Deletes a specified sheet from a workbook and saves it to a new file."""
    final_output_path = _get_safe_path(file_output_dir, output_filename)
    try:
        workbook = load_workbook(filename=file_path)
        if sheet_to_delete not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet_to_delete}' not found.")
        if len(workbook.sheetnames) <= 1:
            raise ValueError("Cannot delete the only sheet in the workbook.")
            
        # Remove the sheet
        del workbook[sheet_to_delete]
        workbook.save(final_output_path)
        workbook.close()
        
        return {"success": True, "output_file": final_output_path, "deleted_sheet_name": sheet_to_delete}
    except Exception as e:
        raise Exception(f"An unexpected error occurred while deleting the sheet: {e}")

def duplicate_sheet(file_path: str, file_output_dir: str, output_filename: str, source_sheet: str, new_sheet_name: str) -> Dict[str, Any]:
    """Duplicates an existing sheet and gives it a new name."""
    final_output_path = _get_safe_path(file_output_dir, output_filename)
    try:
        workbook = load_workbook(filename=file_path)
        if source_sheet not in workbook.sheetnames:
            raise ValueError(f"Source sheet '{source_sheet}' not found.")
        if new_sheet_name in workbook.sheetnames:
            raise ValueError(f"A sheet named '{new_sheet_name}' already exists.")
            
        source_worksheet = workbook[source_sheet]
        new_worksheet = workbook.copy_worksheet(source_worksheet)
        new_worksheet.title = new_sheet_name
        
        workbook.save(final_output_path)
        workbook.close()
        
        return {"success": True, "output_file": final_output_path, "duplicated_from": source_sheet, "duplicated_to": new_sheet_name}
    except Exception as e:
        raise Exception(f"An unexpected error occurred while duplicating the sheet: {e}")

from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

def apply_conditional_formatting(file_path: str, file_output_dir: str, output_filename: str, column_name: str, operator: str, value: Any, color: str, sheet_name: str = None) -> Dict[str, Any]:
    """Applies conditional formatting to a column based on a simple rule (e.g., cell > value)."""
    final_output_path = _get_safe_path(file_output_dir, output_filename)
    try:
        workbook = load_workbook(filename=file_path)
        sheet = workbook[sheet_name] if sheet_name else workbook.active

        # Find the column index
        header = [cell.value for cell in sheet[1]]
        if column_name not in header:
            raise ValueError(f"Column '{column_name}' not found.")
        col_idx_letter = chr(ord('A') + header.index(column_name))

        # Define the fill color
        fills = {
            'red': PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid'),
            'green': PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid'),
            'yellow': PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
        }
        if color not in fills:
            raise ValueError(f"Color '{color}' is not supported. Use red, green, or yellow.")
        fill = fills[color]

        # Define the rule
        rule = CellIsRule(operator=operator, formula=[value], fill=fill)

        # Apply the rule to the entire column (e.g., A2:A1048576)
        range_to_format = f"{col_idx_letter}2:{col_idx_letter}{sheet.max_row}"
        sheet.conditional_formatting.add(range_to_format, rule)

        workbook.save(final_output_path)
        workbook.close()

        return {"success": True, "output_file": final_output_path, "formatted_column": column_name}
    except Exception as e:
        raise Exception(f"An unexpected error occurred while applying conditional formatting: {e}")
